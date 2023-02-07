import os

import openpyxl
import pytest
from openpyxl.styles import Font


@pytest.fixture
def setup_excel():
    value = [
        [8, 31, 59, 66, 23, "test"],
        [69, 57, 76, 66, 23, 90],
        [17, 19, 90, 66, 23, "test"],
        [27, 38, 98, 25, 56, 89],
        [10, 12, 42, 30, 26, 1],
        [12924, 32580, "test", 2249, 91631, "test"],
        [76625, 93205, "test", 30761, 75471, "test"]]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Test1"
    font = Font(name='メイリオ', size=14, bold=True, italic=True, underline='single')
    for row in range(1, 5):
        for col in range(1, 6):
            sheet.cell(row=row, column=col).value = value[row-1][col-1]
    for row in sheet:
        for cell in row:
            sheet[cell.coordinate].font = font
    sheet["H1"].value = 1000
    sheet.merge_cells('H1:I1')
    sheet["A7"].value = "SampleText"
    sheet.merge_cells('A7:G9')
    workbook.save("tests/sample.xlsx")
    yield
    os.remove("tests/sample.xlsx")
    